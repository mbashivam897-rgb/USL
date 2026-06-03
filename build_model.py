"""
Build United Spirits Ltd - PPE, CWIP & ROU schedules (Standalone, Rs crores).
Historicals FY2022-FY2026 (from annual reports / audited results), forecast FY2027-FY2030.
Output: United_Spirits_PPE_CWIP_ROU_Model.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---------- styling helpers ----------
NAVY   = "1F3864"
BLUE   = "2F5496"
LBLUE  = "DDEBF7"   # historical shade
LGREEN = "E2EFDA"   # forecast shade
LGREY  = "F2F2F2"
GOLD   = "FFF2CC"
WHITE  = "FFFFFF"

f_title   = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
f_sub     = Font(name="Calibri", size=9, italic=True, color="FFFFFF")
f_hdr     = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
f_sec     = Font(name="Calibri", size=10, bold=True, color=NAVY)
f_lbl     = Font(name="Calibri", size=10, color="000000")
f_lbl_b   = Font(name="Calibri", size=10, bold=True, color="000000")
f_input   = Font(name="Calibri", size=10, color="0000CC")          # blue = input/hardcode
f_hist    = Font(name="Calibri", size=10, color="000000")          # reported hardcode
f_form    = Font(name="Calibri", size=10, color="000000")          # formula
f_note    = Font(name="Calibri", size=9, italic=True, color="595959")

fill_title = PatternFill("solid", fgColor=NAVY)
fill_hdr   = PatternFill("solid", fgColor=BLUE)
fill_hist  = PatternFill("solid", fgColor=LBLUE)
fill_fcst  = PatternFill("solid", fgColor=LGREEN)
fill_sec   = PatternFill("solid", fgColor=LGREY)
fill_gold  = PatternFill("solid", fgColor=GOLD)
fill_tot   = PatternFill("solid", fgColor="D9E1F2")

thin = Side(style="thin", color="BFBFBF")
med  = Side(style="medium", color="808080")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
border_top = Border(top=Side(style="thin", color="808080"))
border_topbot = Border(top=Side(style="thin", color="808080"), bottom=Side(style="double", color="808080"))

NUM = '#,##0;(#,##0)'
NUM1 = '#,##0.0;(#,##0.0)'
PCT = '0.0%'

# Year columns: C..K  => FY2022..FY2030
YEARS = ["FY2022","FY2023","FY2024","FY2025","FY2026","FY2027E","FY2028E","FY2029E","FY2030E"]
COLS  = ["C","D","E","F","G","H","I","J","K"]
HIST_COLS = ["C","D","E","F","G"]      # FY22-26
FCST_COLS = ["H","I","J","K"]          # FY27-30
col_idx = {y:c for y,c in zip(YEARS,COLS)}

def set_year_header(ws, row, label="(Rs crore)"):
    ws.cell(row=row, column=1, value=label).font = f_hdr
    ws.cell(row=row, column=1).fill = fill_hdr
    ws.cell(row=row, column=2).fill = fill_hdr
    for y,c in zip(YEARS,COLS):
        cell = ws[f"{c}{row}"]
        cell.value = y
        cell.font = f_hdr
        cell.fill = fill_hdr
        cell.alignment = Alignment(horizontal="center")
    # shade hint row beneath handled separately

def shade_cols(ws, row):
    for c in HIST_COLS:
        ws[f"{c}{row}"].fill = fill_hist
    for c in FCST_COLS:
        ws[f"{c}{row}"].fill = fill_fcst

def style_label(ws, row, text, bold=False, indent=1, section=False):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = f_sec if section else (f_lbl_b if bold else f_lbl)
    cell.alignment = Alignment(indent=indent)
    if section:
        for col in range(1, 12):
            ws.cell(row=row, column=col).fill = fill_sec

def put(ws, col, row, value, kind="form", numfmt=NUM):
    cell = ws[f"{col}{row}"]
    cell.value = value
    if kind == "input":
        cell.font = f_input
    elif kind == "hist":
        cell.font = f_hist
    else:
        cell.font = f_form
    cell.number_format = numfmt
    cell.alignment = Alignment(horizontal="right")
    return cell

def widths(ws):
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 7
    for c in COLS:
        ws.column_dimensions[c].width = 11

# =====================================================================
# SHEET 1: COVER & NOTES
# =====================================================================
ws = wb.active
ws.title = "Cover & Notes"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 100

ws.merge_cells("B2:B2")
ws["B2"] = "UNITED SPIRITS LIMITED (Diageo India)"
ws["B2"].font = Font(size=16, bold=True, color=NAVY)
ws["B3"] = "Fixed Asset Schedules - Property, Plant & Equipment (PPE), Capital Work-in-Progress (CWIP) & Right-of-Use (ROU) assets"
ws["B3"].font = Font(size=11, bold=True, color=BLUE)
ws["B4"] = "Standalone basis  |  All figures in Rs crore  |  Financial year ending 31 March"
ws["B4"].font = Font(size=10, italic=True, color="595959")

notes = [
    ("", ""),
    ("SCOPE", "sec"),
    ("Historical actuals: FY2022 to FY2026 (5 years).  Forecast: FY2027E to FY2030E (4 years).", "n"),
    ("Three schedules are built as roll-forwards and linked: cash capex flows into CWIP, is capitalised into", "n"),
    ("PPE, and is depreciated; ROU assets follow a separate lease roll-forward under Ind AS 116.", "n"),
    ("", ""),
    ("HOW TO USE", "sec"),
    ("Blue figures = inputs / reported hardcodes you can change.  Black figures = formulas.", "n"),
    ("Historical columns are shaded blue; forecast columns are shaded green.", "n"),
    ("All forecast drivers live on the 'Assumptions' tab - change them to flex the model.", "n"),
    ("", ""),
    ("DATA SOURCES (primary)", "sec"),
    ("- USL Integrated Annual Report FY2024-25 (Notes 3.1 PPE, 3.2 ROU, 3.3 CWIP) - FY24 & FY25", "n"),
    ("- USL Annual Report FY2023-24 (Notes 3.1/3.2/3.3) - FY23 & FY24", "n"),
    ("- USL Annual Report FY2021-22 (reported in INR millions; converted /10) - FY22", "n"),
    ("- USL Audited financial results for year ended 31 March 2026 - FY26 balance sheet, cash flow", "n"),
    ("- Cross-checked vs screener.in 'Fixed Assets' = PPE + ROU + Intangibles + Investment property", "n"),
    ("", ""),
    ("KEY NOTES & CAVEATS", "sec"),
    ("1) FY2022 is shown on a RESTATED basis (post the Pioneer Distilleries merger) so that it ties to the", "n"),
    ("   FY2023 opening balances. As originally reported (pre-merger) FY22 net PPE was ~Rs 1,016 cr.", "n"),
    ("2) FY2026 detailed PPE/ROU notes are not yet published (annual report pending). FY26 closing", "n"),
    ("   balances, total D&A (Rs 283 cr) and cash capex (Rs 181 cr) are taken from the audited results;", "n"),
    ("   the split of FY26 depreciation / additions between PPE & ROU is estimated (flagged in-sheet).", "n"),
    ("3) The PPE schedule is run on a NET-block roll-forward; a memo shows reported gross block &", "n"),
    ("   accumulated depreciation for FY22r-FY25 as a check.", "n"),
    ("4) Schedules exclude Intangible assets and Investment property (these are separate line items).", "n"),
    ("5) This is an analytical model for illustration, not investment advice.", "n"),
    ("", ""),
    ("TABS", "sec"),
    ("Assumptions  |  PPE Schedule  |  CWIP Schedule  |  ROU Schedule  |  Summary", "n"),
]
r = 6
for text, kind in notes:
    c = ws.cell(row=r, column=2, value=text)
    if kind == "sec":
        c.font = Font(size=10, bold=True, color=NAVY)
        c.fill = PatternFill("solid", fgColor=GOLD)
    elif kind == "n":
        c.font = Font(size=10, color="000000")
    r += 1

# =====================================================================
# SHEET 2: ASSUMPTIONS
# =====================================================================
wsA = wb.create_sheet("Assumptions")
wsA.sheet_view.showGridLines = False
widths(wsA)
wsA.merge_cells("A1:K1")
wsA["A1"] = "ASSUMPTIONS & FORECAST DRIVERS  (edit the blue cells)"
wsA["A1"].font = f_title; wsA["A1"].fill = fill_title
wsA["A1"].alignment = Alignment(horizontal="left", vertical="center")
wsA.row_dimensions[1].height = 22
set_year_header(wsA, 3, "(Rs crore unless stated)")

# Row map
# 4 NSV, 5 growth, 6 capex%, 7 PPE dep%, 8 ROU add%, 9 ROU amort%
style_label(wsA, 4, "Net sales value (NSV) - forecast driver base", bold=True)
style_label(wsA, 5, "  NSV growth (YoY)")
style_label(wsA, 6, "  Cash capex as % of NSV")
style_label(wsA, 7, "  PPE depreciation as % of opening net block")
style_label(wsA, 8, "  ROU additions as % of NSV")
style_label(wsA, 9, "  ROU amortisation as % of (opening + 1/2 additions)")
for rr in (4,5,6,7,8,9):
    shade_cols(wsA, rr)

# NSV: FY26 base input, FY22-25 left as reported approx (we only need FY26 base for forecast)
# Provide FY26 base (sourced consolidated NSV proxy) and project forward.
put(wsA, "G", 4, 12467, "input")   # FY26 NSV base
for c in HIST_COLS[:-1]:           # FY22-FY25 leave blank (not required)
    wsA[f"{c}4"].value = None
# forecast NSV = prior*(1+growth)
for i,c in enumerate(FCST_COLS):
    prev = COLS[COLS.index(c)-1]
    put(wsA, c, 4, f"={prev}4*(1+{c}5)", "form")

# NSV growth (forecast inputs)
for c in FCST_COLS:
    put(wsA, c, 5, 0.09, "input", PCT)
# show FY26 implied n/a
# capex % of NSV
put(wsA, "G", 6, "=181/G4", "form", PCT)   # FY26 implied 1.45%
for c in FCST_COLS:
    put(wsA, c, 6, 0.015, "input", PCT)
# PPE dep % (implied history shown FY24-26 for reference in PPE sheet); forecast input
for c in FCST_COLS:
    put(wsA, c, 7, 0.13, "input", PCT)
# ROU additions % of NSV
for c in FCST_COLS:
    put(wsA, c, 8, 0.013, "input", PCT)
# ROU amort %
for c in FCST_COLS:
    put(wsA, c, 9, 0.30, "input", PCT)

# notes
nrow = 12
for txt in [
    "Notes:",
    "- NSV FY2026 base (Rs 12,467 cr) is USL's reported FY26 net sales value; grown by the YoY% to drive forecast capex & lease additions.",
    "- Cash capex (% of NSV) feeds CWIP additions; CWIP is capitalised into PPE the same year (CWIP held ~flat).",
    "- PPE depreciation % of opening net block is calibrated to FY2024-FY2026 actuals (~12-13%).",
    "- ROU additions (% of NSV) and amortisation (% of opening + half additions) reflect short-life plant/equipment leases.",
    "- All blue cells are editable inputs; change any driver and the schedules + Summary update automatically.",
]:
    c = wsA.cell(row=nrow, column=1, value=txt)
    c.font = f_note if nrow>12 else f_sec
    nrow += 1

# =====================================================================
# SHEET 3: PPE SCHEDULE
# =====================================================================
wsP = wb.create_sheet("PPE Schedule")
wsP.sheet_view.showGridLines = False
widths(wsP)
wsP.merge_cells("A1:K1")
wsP["A1"] = "PROPERTY, PLANT & EQUIPMENT (PPE) SCHEDULE  -  Standalone (Rs crore)"
wsP["A1"].font = f_title; wsP["A1"].fill = fill_title
wsP.row_dimensions[1].height = 22
set_year_header(wsP, 3)

style_label(wsP, 4, "Net block roll-forward", section=True)
style_label(wsP, 5, "Opening net block")
style_label(wsP, 6, "  (+) Additions (capitalised from CWIP)")
style_label(wsP, 7, "  (-) Depreciation")
style_label(wsP, 8, "  (-) Impairment")
style_label(wsP, 9, "  (-) Disposals, transfers & reclassifications (net)")
style_label(wsP, 10, "Closing net block", bold=True)
for rr in range(5,11): shade_cols(wsP, rr)

# FY2022 base: closing hardcode 1201, flows blank
put(wsP, "C", 10, 1201, "hist")
# FY23-FY26 inputs (reported), closing = formula
# Additions
add = {"D":135,"E":126,"F":126,"G":179}
dep = {"D":129,"E":121,"F":113,"G":112}      # G (FY26) PPE dep estimated
imp = {"D":109,"E":20,"F":0,"G":0}
disp= {"D":120,"E":119,"F":7,"G":86}         # net disposals/transfers (derived to tie)
for c in ["D","E","F","G"]:
    prev = COLS[COLS.index(c)-1]
    put(wsP, c, 5, f"={prev}10", "form")           # opening = prior closing
    put(wsP, c, 6, add[c], "hist" if c!="G" else "input")
    put(wsP, c, 7, dep[c], "hist" if c!="G" else "input")
    put(wsP, c, 8, imp[c], "hist")
    put(wsP, c, 9, disp[c], "hist" if c!="G" else "input")
    put(wsP, c, 10, f"={c}5+{c}6-{c}7-{c}8-{c}9", "form")
# Forecast FY27-30
for c in FCST_COLS:
    prev = COLS[COLS.index(c)-1]
    put(wsP, c, 5, f"={prev}10", "form")
    put(wsP, c, 6, f"='CWIP Schedule'!{c}7", "form")        # additions = CWIP capitalised
    put(wsP, c, 7, f"=Assumptions!{c}7*{c}5", "form")       # dep = %*opening
    put(wsP, c, 8, 0, "input")
    put(wsP, c, 9, 0, "input")
    put(wsP, c, 10, f"={c}5+{c}6-{c}7-{c}8-{c}9", "form")

# Memo: reported gross block & acc dep (FY22r-FY25)
style_label(wsP, 12, "Memo: reported gross block & accumulated depreciation", section=True)
style_label(wsP, 13, "Gross block (closing)")
style_label(wsP, 14, "Accumulated depreciation & impairment (closing)")
style_label(wsP, 15, "Net block (check = gross - acc. dep.)", bold=True)
gb = {"C":2202,"D":1988,"E":1869,"F":1915}
ad = {"C":1001,"D":1010,"E":1025,"F":1065}
for c in ["C","D","E","F"]:
    put(wsP, c, 13, gb[c], "hist")
    put(wsP, c, 14, ad[c], "hist")
    put(wsP, c, 15, f"={c}13-{c}14", "form")
wsP["G13"] = "n/a"; wsP["G13"].font = f_note; wsP["G13"].alignment=Alignment(horizontal="right")
wsP["G14"] = "n/a"; wsP["G14"].font = f_note; wsP["G14"].alignment=Alignment(horizontal="right")

# checks
style_label(wsP, 17, "Checks", section=True)
style_label(wsP, 18, "  Implied depreciation % of opening net block")
for c in ["D","E","F","G"]+FCST_COLS:
    put(wsP, c, 18, f"={c}7/{c}5", "form", PCT)
style_label(wsP, 19, "  Net block memo vs roll-forward (should be 0)")
for c in ["C","D","E","F"]:
    put(wsP, c, 19, f"={c}15-{c}10", "form")

for txt,row in [("FY26 PPE depreciation & disposal/transfer split are estimates (full FY26 notes pending).",20)]:
    cc=wsP.cell(row=row,column=1,value=txt); cc.font=f_note

# =====================================================================
# SHEET 4: CWIP SCHEDULE
# =====================================================================
wsC = wb.create_sheet("CWIP Schedule")
wsC.sheet_view.showGridLines = False
widths(wsC)
wsC.merge_cells("A1:K1")
wsC["A1"] = "CAPITAL WORK-IN-PROGRESS (CWIP) SCHEDULE  -  Standalone (Rs crore)"
wsC["A1"].font = f_title; wsC["A1"].fill = fill_title
wsC.row_dimensions[1].height = 22
set_year_header(wsC, 3)
style_label(wsC, 4, "CWIP roll-forward", section=True)
style_label(wsC, 5, "Opening CWIP")
style_label(wsC, 6, "  (+) Additions (capital expenditure incurred)")
style_label(wsC, 7, "  (-) Capitalised to PPE / intangibles")
style_label(wsC, 8, "Closing CWIP", bold=True)
for rr in range(5,9): shade_cols(wsC, rr)
# FY22 base
put(wsC, "C", 8, 88, "hist")
cadd = {"D":114,"E":96,"F":161,"G":179}
ccap = {"D":135,"E":126,"F":126,"G":181}
for c in ["D","E","F","G"]:
    prev = COLS[COLS.index(c)-1]
    put(wsC, c, 5, f"={prev}8", "form")
    put(wsC, c, 6, cadd[c], "hist" if c!="G" else "input")
    put(wsC, c, 7, ccap[c], "hist" if c!="G" else "input")
    put(wsC, c, 8, f"={c}5+{c}6-{c}7", "form")
# forecast: additions = NSV*capex%; capitalised = additions (flat CWIP)
for c in FCST_COLS:
    prev = COLS[COLS.index(c)-1]
    put(wsC, c, 5, f"={prev}8", "form")
    put(wsC, c, 6, f"=Assumptions!{c}4*Assumptions!{c}6", "form")
    put(wsC, c, 7, f"={c}6", "form")
    put(wsC, c, 8, f"={c}5+{c}6-{c}7", "form")
style_label(wsC, 10, "Checks", section=True)
style_label(wsC, 11, "  Cash capex (= additions)")
for c in ["D","E","F","G"]+FCST_COLS:
    put(wsC, c, 11, f"={c}6", "form")
cc=wsC.cell(row=12,column=1,value="FY26 additions/capitalised tie to reported cash capex (~Rs 181 cr); CWIP held ~flat in forecast.")
cc.font=f_note

# =====================================================================
# SHEET 5: ROU SCHEDULE
# =====================================================================
wsR = wb.create_sheet("ROU Schedule")
wsR.sheet_view.showGridLines = False
widths(wsR)
wsR.merge_cells("A1:K1")
wsR["A1"] = "RIGHT-OF-USE (ROU) ASSETS SCHEDULE  -  Standalone, Ind AS 116 (Rs crore)"
wsR["A1"].font = f_title; wsR["A1"].fill = fill_title
wsR.row_dimensions[1].height = 22
set_year_header(wsR, 3)
style_label(wsR, 4, "ROU asset roll-forward", section=True)
style_label(wsR, 5, "Opening ROU assets")
style_label(wsR, 6, "  (+) Additions & remeasurements")
style_label(wsR, 7, "  (-) Amortisation (depreciation of ROU)")
style_label(wsR, 8, "  (-) Terminations, transfers & other (net)")
style_label(wsR, 9, "Closing ROU assets", bold=True)
for rr in range(5,10): shade_cols(wsR, rr)
put(wsR, "C", 9, 261, "hist")
radd = {"D":86,"E":184,"F":394,"G":82}
ramo = {"D":130,"E":129,"F":147,"G":158}
rter = {"D":44,"E":1,"F":17,"G":0}
for c in ["D","E","F","G"]:
    prev = COLS[COLS.index(c)-1]
    put(wsR, c, 5, f"={prev}9", "form")
    put(wsR, c, 6, radd[c], "hist" if c!="G" else "input")
    put(wsR, c, 7, ramo[c], "hist" if c!="G" else "input")
    put(wsR, c, 8, rter[c], "hist")
    put(wsR, c, 9, f"={c}5+{c}6-{c}7-{c}8", "form")
# forecast: additions = NSV*ROUadd%; amort = ROUamort%*(open+0.5*add)
for c in FCST_COLS:
    prev = COLS[COLS.index(c)-1]
    put(wsR, c, 5, f"={prev}9", "form")
    put(wsR, c, 6, f"=Assumptions!{c}4*Assumptions!{c}8", "form")
    put(wsR, c, 7, f"=Assumptions!{c}9*({c}5+0.5*{c}6)", "form")
    put(wsR, c, 8, 0, "input")
    put(wsR, c, 9, f"={c}5+{c}6-{c}7-{c}8", "form")

# memo lease economics (where available)
style_label(wsR, 11, "Memo: lease economics (reported)", section=True)
style_label(wsR, 12, "  Lease liabilities (closing)")
style_label(wsR, 13, "  Interest on lease liabilities (P&L)")
style_label(wsR, 14, "  Principal repayment of lease liabilities (CF)")
llia = {"C":264,"D":182,"E":240,"F":480}
lint = {"E":21,"F":40,"G":37}
lpri = {"E":126,"F":137,"G":143}
for c,v in llia.items(): put(wsR, c, 12, v, "hist")
for c,v in lint.items(): put(wsR, c, 13, v, "hist")
for c,v in lpri.items(): put(wsR, c, 14, v, "hist")
style_label(wsR, 16, "Checks", section=True)
style_label(wsR, 17, "  Implied ROU amortisation % of opening")
for c in ["D","E","F","G"]+FCST_COLS:
    put(wsR, c, 17, f"={c}7/{c}5", "form", PCT)
cc=wsR.cell(row=18,column=1,value="FY26 ROU additions/amortisation are estimates (full FY26 lease note pending); closing Rs 381 cr is reported.")
cc.font=f_note

# =====================================================================
# SHEET 6: SUMMARY
# =====================================================================
wsS = wb.create_sheet("Summary")
wsS.sheet_view.showGridLines = False
widths(wsS)
wsS.merge_cells("A1:K1")
wsS["A1"] = "SUMMARY  -  PPE, CWIP & ROU  |  United Spirits (Standalone, Rs crore)"
wsS["A1"].font = f_title; wsS["A1"].fill = fill_title
wsS.row_dimensions[1].height = 22
set_year_header(wsS, 3)
style_label(wsS, 4, "Closing balances (Balance Sheet)", section=True)
style_label(wsS, 5, "Property, plant & equipment (net)")
style_label(wsS, 6, "Right-of-use assets")
style_label(wsS, 7, "Capital work-in-progress")
style_label(wsS, 8, "Total PPE + ROU + CWIP", bold=True)
for rr in range(5,9): shade_cols(wsS, rr)
for c in COLS:
    put(wsS, c, 5, f"='PPE Schedule'!{c}10", "form")
    put(wsS, c, 6, f"='ROU Schedule'!{c}9", "form")
    put(wsS, c, 7, f"='CWIP Schedule'!{c}8", "form")
    cell = put(wsS, c, 8, f"=SUM({c}5:{c}7)", "form")
    cell.font = f_lbl_b
    cell.fill = fill_tot

style_label(wsS, 10, "Flows (P&L / Cash flow linkage)", section=True)
style_label(wsS, 11, "Cash capex (PPE/CWIP)")
style_label(wsS, 12, "PPE depreciation")
style_label(wsS, 13, "PPE impairment")
style_label(wsS, 14, "ROU amortisation")
style_label(wsS, 15, "Total depreciation & amortisation (PPE + ROU)", bold=True)
for rr in range(11,16): shade_cols(wsS, rr)
for c in COLS:
    put(wsS, c, 11, f"='CWIP Schedule'!{c}6", "form")
    put(wsS, c, 12, f"='PPE Schedule'!{c}7", "form")
    put(wsS, c, 13, f"='PPE Schedule'!{c}8", "form")
    put(wsS, c, 14, f"='ROU Schedule'!{c}7", "form")
    cell = put(wsS, c, 15, f"={c}12+{c}14", "form")
    cell.font = f_lbl_b; cell.fill = fill_tot
# FY22 flows blank (base) -> clear cap/dep refs that point to blank become 0; acceptable

style_label(wsS, 17, "Growth & ratios", section=True)
style_label(wsS, 18, "  Total fixed-asset base YoY growth")
for c in COLS[1:]:
    prev = COLS[COLS.index(c)-1]
    put(wsS, c, 18, f"=IFERROR({c}8/{prev}8-1,\"\")", "form", PCT)

# freeze panes on schedule sheets
for w in (wsA,wsP,wsC,wsR,wsS):
    w.freeze_panes = "C4"

wb.save("United_Spirits_PPE_CWIP_ROU_Model.xlsx")
print("Saved United_Spirits_PPE_CWIP_ROU_Model.xlsx")
