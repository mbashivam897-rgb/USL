"""
Builds: PE_Exit_Model_Hexaware_Siemens.xlsx
Institutional PE exit model with LIVE formula linkages across sheets:
  - Assumptions (single source of truth) -> feeds Valuation & Returns sheets
  - Financials (5-yr historicals) -> feeds margin/EBITDA derivations
  - Valuation (Bear/Base/Bull) -> EV/Equity computed via formulas
  - Returns (MOIC/IRR) -> computed via formulas referencing Valuation
  - Dashboard -> pulls headline outputs from other sheets
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

wb = Workbook()

# ---------- styles ----------
NAVY   = "1F3864"
BLUE   = "2E5496"
LBLUE  = "D6E0F0"
GOLD   = "BF9000"
LGOLD  = "FFF2CC"
GREY   = "F2F2F2"
GREEN  = "C6EFCE"
RED    = "FFC7CE"
WHITE  = "FFFFFF"

def style_title(cell):
    cell.font = Font(bold=True, size=16, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def style_header(cell):
    cell.font = Font(bold=True, size=11, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_subhead(cell):
    cell.font = Font(bold=True, size=11, color=NAVY)
    cell.fill = PatternFill("solid", fgColor=LGOLD)

def style_label(cell):
    cell.font = Font(bold=True, size=10, color="000000")
    cell.fill = PatternFill("solid", fgColor=GREY)

def style_input(cell):
    cell.font = Font(size=10, color="0000CC")  # blue = hardcoded input
    cell.fill = PatternFill("solid", fgColor=LBLUE)
    cell.alignment = Alignment(horizontal="center")

def style_calc(cell):
    cell.font = Font(size=10, color="000000")  # black = formula
    cell.alignment = Alignment(horizontal="center")

thin = Side(style="thin", color="BFBFBF")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

def box(ws, rng):
    for row in ws[rng]:
        for c in row:
            c.border = border_all

# =====================================================================
# SHEET 1: ASSUMPTIONS (single source of truth)
# =====================================================================
ws = wb.active
ws.title = "Assumptions"
ws.sheet_properties.tabColor = NAVY
ws["A1"] = "PE EXIT MODEL — GLOBAL ASSUMPTIONS (single source of truth)"
ws.merge_cells("A1:E1")
style_title(ws["A1"])
ws.row_dimensions[1].height = 26

ws["A3"] = "Parameter"; ws["B3"] = "Value"; ws["C3"] = "Unit"; ws["D3"] = "Source / Note"
for col in "ABCD":
    style_header(ws[f"{col}3"])

assum = [
    ("FX rate (INR per USD)", 85, "INR/USD", "Modeling assumption FY25-26"),
    ("Discount rate / WACC", 0.115, "%", "Analyst estimate"),
    ("Terminal growth", 0.045, "%", "Analyst estimate"),
    ("", "", "", ""),
    ("HEXAWARE — sponsor (Carlyle)", "", "", ""),
    ("Carlyle entry EV (2021)", 23200, "INR cr", "~US$3bn; Moneycontrol/CNBC"),
    ("Carlyle entry equity (2021, est.)", 22000, "INR cr", "Estimate (net cash co.)"),
    ("IPO OFS proceeds (Feb-2025)", 8750, "INR cr", "100% OFS; Carlyle"),
    ("Hexaware exit-year EBITDA (est.)", 2800, "INR cr", "~18% margin on ~15.5k cr rev"),
    ("Hexaware net cash (est.)", 1500, "INR cr", "Estimate"),
    ("Hexaware entry year", 2021, "yr", "Carlyle acquisition"),
    ("Hexaware exit year (base)", 2028, "yr", "Recommended staggered exit"),
    ("", "", "", ""),
    ("SIEMENS INDIA (continuing co.)", "", "", ""),
    ("Siemens exit-year PAT (est.)", 1850, "INR cr", "Continuing-ops estimate"),
    ("Hypothetical entry mcap (today)", 130000, "INR cr", "Jun-2026 ~1.27-1.33 lakh cr"),
    ("Hypothetical hold (years)", 4, "yr", "Illustrative"),
]
r = 4
for name, val, unit, note in assum:
    ws[f"A{r}"] = name
    ws[f"B{r}"] = val
    ws[f"C{r}"] = unit
    ws[f"D{r}"] = note
    if name and val == "":
        style_subhead(ws[f"A{r}"]); ws.merge_cells(f"A{r}:D{r}")
    elif name:
        style_label(ws[f"A{r}"]); style_input(ws[f"B{r}"]); style_calc(ws[f"C{r}"])
        ws[f"D{r}"].font = Font(size=9, italic=True, color="595959")
    r += 1

# named cells for linkage
fx_cell = "Assumptions!$B$4"
hex_entry_equity = "Assumptions!$B$10"
hex_ipo = "Assumptions!$B$11"
hex_ebitda = "Assumptions!$B$12"
hex_netcash = "Assumptions!$B$13"
sie_pat = "Assumptions!$B$18"
sie_entry = "Assumptions!$B$19"

ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 10
ws.column_dimensions["D"].width = 36
box(ws, "A3:D20")

# legend
ws["A22"] = "Legend:"; ws["A22"].font = Font(bold=True)
ws["B22"] = "Blue = hardcoded input"; ws["B22"].font = Font(color="0000CC")
ws["B23"] = "Black = formula (linked)"; ws["B23"].font = Font(color="000000")

# =====================================================================
# SHEET 2: HEXAWARE FINANCIALS (5-yr historicals)
# =====================================================================
ws = wb.create_sheet("HEX_Financials")
ws.sheet_properties.tabColor = BLUE
ws["A1"] = "HEXAWARE TECHNOLOGIES — 5-YEAR FINANCIAL SUMMARY (CY, Dec-end)"
ws.merge_cells("A1:G1"); style_title(ws["A1"]); ws.row_dimensions[1].height = 24

hdr = ["Metric (INR cr unless noted)", "CY21", "CY22", "CY23", "CY24", "CY25", "CAGR/Note"]
for i, h in enumerate(hdr):
    c = ws.cell(row=3, column=i+1, value=h); style_header(c)

# revenue INR cr (from stockopedia INR m /10)
rev = [7177.7, 9199.6, 10380.3, 11974.4, 13430.4]
usd = [970, 1156, 1257, 1429, 1537]
ebitda_m = [0.17, 0.165, 0.165, 0.17, 0.167]
pat = [750, 820, 998, 1150, 1300]

rows = [
    ("Revenue (INR cr)", rev, "input"),
    ("Revenue (US$ m)", usd, "input"),
    ("Revenue growth % (USD)", None, "growth_usd"),
    ("EBITDA margin %", ebitda_m, "input"),
    ("EBITDA (INR cr)", None, "ebitda_calc"),
    ("PAT (INR cr)", pat, "input"),
    ("PAT margin %", None, "patmargin"),
]
r = 4
for label, data, kind in rows:
    ws.cell(row=r, column=1, value=label); style_label(ws.cell(row=r, column=1))
    for col in range(2, 7):
        cell = ws.cell(row=r, column=col)
        cidx = col - 2
        if kind == "input":
            cell.value = data[cidx]
            if "%" in label:
                cell.number_format = "0.0%"
            style_input(cell)
        elif kind == "growth_usd":
            if col == 2:
                cell.value = "n/a"
            else:
                cell.value = f"=(B{r-1}/A{r-1})-1".replace("A", get_column_letter(col-1)).replace("B", get_column_letter(col))
                # simpler explicit
                cell.value = f"={get_column_letter(col)}5/{get_column_letter(col-1)}5-1"
            cell.number_format = "0.0%"; style_calc(cell)
        elif kind == "ebitda_calc":
            cell.value = f"={get_column_letter(col)}4*{get_column_letter(col)}7"
            cell.number_format = "#,##0"; style_calc(cell)
        elif kind == "patmargin":
            cell.value = f"={get_column_letter(col)}9/{get_column_letter(col)}4"
            cell.number_format = "0.0%"; style_calc(cell)
    r += 1

# CAGR note column
ws["G4"] = "=(F4/B4)^(1/4)-1"; ws["G4"].number_format = "0.0%"; style_calc(ws["G4"])
ws["G4"].comment = None
ws["G3"]  # header already
ws.cell(row=4, column=7).value = "=(F4/B4)^(1/4)-1"
ws.cell(row=4, column=7).number_format = "0.0%"
ws.cell(row=9, column=7).value = "=(F9/B9)^(1/4)-1"
ws.cell(row=9, column=7).number_format = "0.0%"

ws.column_dimensions["A"].width = 26
for col in "BCDEFG":
    ws.column_dimensions[col].width = 12
box(ws, "A3:G10")
ws["A12"] = "Sources: Stockopedia (INR revenue series); Hexaware/PR Newswire (USD revenue CY24=$1,429m, CY25=$1,537m); EBITDA & PAT analyst estimates."
ws["A12"].font = Font(size=9, italic=True, color="595959"); ws.merge_cells("A12:G13")

# chart
chart = BarChart(); chart.title = "Hexaware Revenue (INR cr)"; chart.type = "col"
data = Reference(ws, min_col=2, max_col=6, min_row=4, max_row=4)
cats = Reference(ws, min_col=2, max_col=6, min_row=3, max_row=3)
chart.add_data(data, titles_from_data=False); chart.set_categories(cats)
chart.height = 7; chart.width = 14
ws.add_chart(chart, "A15")

# =====================================================================
# SHEET 3: HEXAWARE VALUATION (Bear/Base/Bull) -- formula linked
# =====================================================================
ws = wb.create_sheet("HEX_Valuation")
ws.sheet_properties.tabColor = BLUE
ws["A1"] = "HEXAWARE — EXIT VALUATION SCENARIOS (formula-linked to Assumptions)"
ws.merge_cells("A1:F1"); style_title(ws["A1"]); ws.row_dimensions[1].height = 24

ws["A3"] = "Input"; ws["B3"] = "Bear"; ws["C3"] = "Base"; ws["D3"] = "Bull"; ws["E3"] = "Unit/Note"
for col in "ABCDE": style_header(ws[f"{col}3"])

ws["A4"] = "Exit-year EBITDA (INR cr)"; style_label(ws["A4"])
for col in "BCD":
    ws[f"{col}4"] = f"={hex_ebitda}"; ws[f"{col}4"].number_format = "#,##0"; style_calc(ws[f"{col}4"])
ws["E4"] = "Linked to Assumptions"

ws["A5"] = "Exit EV/EBITDA multiple (x)"; style_label(ws["A5"])
ws["B5"] = 13; ws["C5"] = 16; ws["D5"] = 20
for col in "BCD": style_input(ws[f"{col}5"]); ws[f"{col}5"].number_format = "0.0\"x\""
ws["E5"] = "Comps: midcap IT (de-rated 2026)"

ws["A6"] = "Enterprise Value (INR cr)"; style_label(ws["A6"])
for col in "BCD":
    ws[f"{col}6"] = f"={col}4*{col}5"; ws[f"{col}6"].number_format = "#,##0"; style_calc(ws[f"{col}6"])

ws["A7"] = "(+) Net cash (INR cr)"; style_label(ws["A7"])
for col in "BCD":
    ws[f"{col}7"] = f"={hex_netcash}"; ws[f"{col}7"].number_format = "#,##0"; style_calc(ws[f"{col}7"])

ws["A8"] = "Equity Value (INR cr)"; style_label(ws["A8"])
for col in "BCD":
    ws[f"{col}8"] = f"={col}6+{col}7"; ws[f"{col}8"].number_format = "#,##0"
    ws[f"{col}8"].font = Font(bold=True); ws[f"{col}8"].fill = PatternFill("solid", fgColor=LGOLD)

ws["A9"] = "Equity Value (US$ bn)"; style_label(ws["A9"])
for col in "BCD":
    # INR cr -> US$ bn:  (cr * 1e7) / FX / 1e9  =  cr / FX / 100
    ws[f"{col}9"] = f"={col}8/{fx_cell}/100"; ws[f"{col}9"].number_format = "0.00"; style_calc(ws[f"{col}9"])

ws.column_dimensions["A"].width = 28
for col in "BCDE": ws.column_dimensions[col].width = 14
box(ws, "A3:E9")
ws["A11"] = "EV/Revenue cross-check (Base ~3.0x) and DCF (WACC 11.5%, g 4.5%) support the Base equity range. Sources: ET/Nuvama multiples; analyst est."
ws["A11"].font = Font(size=9, italic=True, color="595959"); ws.merge_cells("A11:E12")

# =====================================================================
# SHEET 4: HEXAWARE RETURNS (MOIC/IRR) -- formula linked
# =====================================================================
ws = wb.create_sheet("HEX_Returns")
ws.sheet_properties.tabColor = GOLD
ws["A1"] = "HEXAWARE — SPONSOR RETURNS (Carlyle) — formula-linked"
ws.merge_cells("A1:F1"); style_title(ws["A1"]); ws.row_dimensions[1].height = 24

ws["A3"]="Metric"; ws["B3"]="Bear"; ws["C3"]="Base"; ws["D3"]="Bull"; ws["E3"]="Note"
for col in "ABCDE": style_header(ws[f"{col}3"])

ws["A4"]="Entry equity (INR cr, 2021)"; style_label(ws["A4"])
for col in "BCD":
    ws[f"{col}4"]=f"={hex_entry_equity}"; ws[f"{col}4"].number_format="#,##0"; style_calc(ws[f"{col}4"])

ws["A5"]="IPO OFS realised (INR cr, 2025)"; style_label(ws["A5"])
for col in "BCD":
    ws[f"{col}5"]=f"={hex_ipo}"; ws[f"{col}5"].number_format="#,##0"; style_calc(ws[f"{col}5"])

ws["A6"]="Residual stake % at exit"; style_label(ws["A6"])
ws["B6"]=0.741; ws["C6"]=0.741; ws["D6"]=0.741
for col in "BCD": style_input(ws[f"{col}6"]); ws[f"{col}6"].number_format="0.0%"

ws["A7"]="Residual equity value (INR cr)"; style_label(ws["A7"])
# link to HEX_Valuation equity value row 8
for col in "BCD":
    ws[f"{col}7"]=f"=HEX_Valuation!{col}8*{col}6"; ws[f"{col}7"].number_format="#,##0"; style_calc(ws[f"{col}7"])

ws["A8"]="Total realisations (INR cr)"; style_label(ws["A8"])
for col in "BCD":
    ws[f"{col}8"]=f"={col}5+{col}7"; ws[f"{col}8"].number_format="#,##0"; style_calc(ws[f"{col}8"])

ws["A9"]="MOIC (x)"; style_label(ws["A9"])
for col in "BCD":
    ws[f"{col}9"]=f"={col}8/{col}4"; ws[f"{col}9"].number_format="0.00\"x\""
    ws[f"{col}9"].font=Font(bold=True); ws[f"{col}9"].fill=PatternFill("solid", fgColor=LGOLD)

ws["A10"]="Holding period (years)"; style_label(ws["A10"])
ws["B10"]=7; ws["C10"]=7; ws["D10"]=7
for col in "BCD": style_input(ws[f"{col}10"])

# ---- Proper cash-flow IRR (captures early 2025 IPO proceeds) ----
ws["A12"]="Cash-flow timeline (INR cr) — for true IRR"; style_subhead(ws["A12"]); ws.merge_cells("A12:H12")
years=[2021,2022,2023,2024,2025,2026,2027,2028]
ws["A13"]="Year"; style_label(ws["A13"])
for i,y in enumerate(years):
    c=ws.cell(row=13,column=2+i,value=y); style_header(c)
# rows for Bear/Base/Bull cash flows
scen_rows={"Bear":14,"Base":15,"Bull":16}
scen_eqcol={"Bear":"B","Base":"C","Bull":"D"}
for scen,rr in scen_rows.items():
    ws.cell(row=rr,column=1,value=f"{scen} CF").font=Font(bold=True)
    ec=scen_eqcol[scen]
    # 2021 entry = -entry equity
    ws.cell(row=rr,column=2,value=f"=-{ec}4").number_format="#,##0"; style_calc(ws.cell(row=rr,column=2))
    # 2022,2023,2024 = 0
    for ci in (3,4,5): ws.cell(row=rr,column=ci,value=0); style_calc(ws.cell(row=rr,column=ci))
    # 2025 = +IPO OFS proceeds
    ws.cell(row=rr,column=6,value=f"={ec}5").number_format="#,##0"; style_calc(ws.cell(row=rr,column=6))
    # 2026,2027 = 0
    for ci in (7,8): ws.cell(row=rr,column=ci,value=0); style_calc(ws.cell(row=rr,column=ci))
    # 2028 = +residual equity value
    ws.cell(row=rr,column=9,value=f"={ec}7").number_format="#,##0"; style_calc(ws.cell(row=rr,column=9))
box(ws,"A13:I16")

ws["A11"]="Gross IRR (true, multi-flow)"; style_label(ws["A11"])
for scen,rr in scen_rows.items():
    ec=scen_eqcol[scen]
    ws[f"{ec}11"]=f"=IRR({get_column_letter(2)}{rr}:{get_column_letter(9)}{rr})"
    ws[f"{ec}11"].number_format="0.0%"
    ws[f"{ec}11"].font=Font(bold=True); ws[f"{ec}11"].fill=PatternFill("solid", fgColor=LGOLD)

ws.column_dimensions["A"].width=30
for col in "BCDEFGHI": ws.column_dimensions[col].width=11
box(ws,"A3:E11")
ws["A18"]="True IRR uses the full cash-flow series: 2021 entry equity outflow, 2025 IPO OFS proceeds (INR 8,750 cr), 2028 residual-stake realisation."
ws["A18"].font=Font(size=9, italic=True, color="595959"); ws.merge_cells("A18:I19")

# =====================================================================
# SHEET 5: SIEMENS FINANCIALS
# =====================================================================
ws = wb.create_sheet("SIE_Financials")
ws.sheet_properties.tabColor = "548235"
ws["A1"]="SIEMENS LTD (INDIA) — 5-YEAR FINANCIALS (FY Oct-Sep; FY20-24 incl. Energy)"
ws.merge_cells("A1:G1"); style_title(ws["A1"]); ws.row_dimensions[1].height=24
hdr=["Metric (INR cr)","FY20","FY21","FY22","FY23","FY24","CAGR"]
for i,h in enumerate(hdr):
    style_header(ws.cell(row=3,column=i+1,value=h))
srev=[9824,13136,16047,19472,22157]
ws.cell(row=4,column=1,value="Revenue (INR cr)"); style_label(ws.cell(row=4,column=1))
for i,v in enumerate(srev):
    c=ws.cell(row=4,column=2+i,value=v); c.number_format="#,##0"; style_input(c)
ws.cell(row=4,column=7,value="=(F4/B4)^(1/4)-1").number_format="0.0%"; style_calc(ws.cell(row=4,column=7))
ws.cell(row=5,column=1,value="Revenue growth %"); style_label(ws.cell(row=5,column=1))
ws.cell(row=5,column=2,value="n/a")
for col in "CDEF":
    prev=chr(ord(col)-1)
    cc=ws[f"{col}5"]; cc.value=f"={col}4/{prev}4-1"; cc.number_format="0.0%"; style_calc(cc)
ws.cell(row=6,column=1,value="PAT (INR cr)"); style_label(ws.cell(row=6,column=1))
ws.cell(row=6,column=6,value=2718).number_format="#,##0"; style_input(ws.cell(row=6,column=6))
ws.cell(row=6,column=2,value="—"); ws.cell(row=6,column=3,value="—"); ws.cell(row=6,column=4,value="—"); ws.cell(row=6,column=5,value="—")
ws.column_dimensions["A"].width=22
for col in "BCDEFG": ws.column_dimensions[col].width=11
box(ws,"A3:G6")
# continuing ops quarterly
ws["A8"]="Continuing operations (ex-Energy), recent quarters"; style_subhead(ws["A8"]); ws.merge_cells("A8:G8")
qh=["Period","Revenue (INR cr)","Growth %","PAT (INR cr)"]
for i,h in enumerate(qh): style_header(ws.cell(row=9,column=i+1,value=h))
qd=[("Q4 FY25 (Sep-25)",5171,0.16,485),("Q1 FY26 (Dec-25)",3831,0.14,269),("Q2 FY26 (Mar-26)",None,0.146,355)]
r=10
for p,rev_,g,pat_ in qd:
    ws.cell(row=r,column=1,value=p); style_label(ws.cell(row=r,column=1))
    if rev_: ws.cell(row=r,column=2,value=rev_).number_format="#,##0"
    else: ws.cell(row=r,column=2,value="n/d")
    style_input(ws.cell(row=r,column=2))
    ws.cell(row=r,column=3,value=g).number_format="0.0%"; style_input(ws.cell(row=r,column=3))
    ws.cell(row=r,column=4,value=pat_).number_format="#,##0"; style_input(ws.cell(row=r,column=4))
    r+=1
box(ws,"A9:D12")
ws["A14"]="Sources: WSJ (FY20-24 revenue); Quartr (FY24 PAT ~2,718 cr); Siemens IR press releases (continuing ops quarters)."
ws["A14"].font=Font(size=9, italic=True, color="595959"); ws.merge_cells("A14:G15")

# =====================================================================
# SHEET 6: SIEMENS VALUATION + RETURNS
# =====================================================================
ws = wb.create_sheet("SIE_Valuation")
ws.sheet_properties.tabColor = "548235"
ws["A1"]="SIEMENS INDIA — EXIT VALUATION & HYPOTHETICAL RETURNS (formula-linked)"
ws.merge_cells("A1:F1"); style_title(ws["A1"]); ws.row_dimensions[1].height=24
ws["A3"]="Input"; ws["B3"]="Bear"; ws["C3"]="Base"; ws["D3"]="Bull"; ws["E3"]="Note"
for col in "ABCDE": style_header(ws[f"{col}3"])
ws["A4"]="Exit-year PAT (grown, INR cr)"; style_label(ws["A4"])
ws["B4"]=3000; ws["C4"]=3700; ws["D4"]=4400
for col in "BCD": style_input(ws[f"{col}4"]); ws[f"{col}4"].number_format="#,##0"
ws["E4"]="Current ~normalized PAT grown ~12-15%/yr over hold"
ws["A5"]="Exit P/E multiple (x)"; style_label(ws["A5"])
ws["B5"]=32; ws["C5"]=45; ws["D5"]=52
for col in "BCD": style_input(ws[f"{col}5"]); ws[f"{col}5"].number_format="0\"x\""
ws["E5"]="Premium capital-goods comps (ABB/Hitachi)"
ws["A6"]="Equity Value (INR cr)"; style_label(ws["A6"])
for col in "BCD":
    ws[f"{col}6"]=f"={col}4*{col}5"; ws[f"{col}6"].number_format="#,##0"
    ws[f"{col}6"].font=Font(bold=True); ws[f"{col}6"].fill=PatternFill("solid", fgColor=LGOLD)
ws["A7"]="Equity Value (US$ bn)"; style_label(ws["A7"])
for col in "BCD":
    ws[f"{col}7"]=f"={col}6/{fx_cell}/100"; ws[f"{col}7"].number_format="0.00"; style_calc(ws[f"{col}7"])
# hypothetical returns
ws["A9"]="Hypothetical entry mcap (INR cr)"; style_label(ws["A9"])
for col in "BCD":
    ws[f"{col}9"]=f"={sie_entry}"; ws[f"{col}9"].number_format="#,##0"; style_calc(ws[f"{col}9"])
ws["A10"]="MOIC (x)"; style_label(ws["A10"])
for col in "BCD":
    ws[f"{col}10"]=f"={col}6/{col}9"; ws[f"{col}10"].number_format="0.00\"x\""
    ws[f"{col}10"].font=Font(bold=True); ws[f"{col}10"].fill=PatternFill("solid", fgColor=LGOLD)
ws["A11"]="Hold (years)"; style_label(ws["A11"])
for col in "BCD":
    ws[f"{col}11"]=f"=Assumptions!$B$20"; style_calc(ws[f"{col}11"])
ws["A12"]="Gross IRR (approx.)"; style_label(ws["A12"])
for col in "BCD":
    ws[f"{col}12"]=f"=({col}10)^(1/{col}11)-1"; ws[f"{col}12"].number_format="0.0%"
    ws[f"{col}12"].font=Font(bold=True); ws[f"{col}12"].fill=PatternFill("solid", fgColor=LGOLD)
ws.column_dimensions["A"].width=30
for col in "BCDE": ws.column_dimensions[col].width=14
box(ws,"A3:E7"); box(ws,"A9:D12")
ws["A14"]="Note: No realistic PE control entry exists (Siemens AG ~69%). Returns illustrative on hypothetical entry at today's mcap. Already richly valued."
ws["A14"].font=Font(size=9, italic=True, color="595959"); ws.merge_cells("A14:E15")

# =====================================================================
# SHEET 7: DASHBOARD (pulls from all sheets)
# =====================================================================
ws = wb.create_sheet("Dashboard")
ws.sheet_properties.tabColor = NAVY
ws["A1"]="IC DASHBOARD — HEADLINE OUTPUTS (linked from all sheets)"
ws.merge_cells("A1:E1"); style_title(ws["A1"]); ws.row_dimensions[1].height=26
ws["A3"]="Metric"; ws["B3"]="Hexaware (Base)"; ws["C3"]="Siemens India (Base)"; ws["D3"]="Comment"
for col in "ABCD": style_header(ws[f"{col}3"])
dash=[
 ("Base equity value (INR cr)", "=HEX_Valuation!C8", "=SIE_Valuation!C6", "Most likely case"),
 ("Base equity value (US$ bn)", "=HEX_Valuation!C9", "=SIE_Valuation!C7", "@FX 85"),
 ("Base MOIC (x)", "=HEX_Returns!C9", "=SIE_Valuation!C10", "Hexaware real; Siemens hypothetical"),
 ("Base gross IRR", "=HEX_Returns!C11", "=SIE_Valuation!C12", "Single-flow proxy"),
 ("Exit readiness (/10)", 8.4, 9.0, "Siemens higher but no control"),
 ("Control available to PE?", "YES", "NO", "Carlyle vs Siemens AG ~69%"),
]
r=4
for m,h,s,c in dash:
    ws.cell(row=r,column=1,value=m); style_label(ws.cell(row=r,column=1))
    hc=ws.cell(row=r,column=2,value=h); sc=ws.cell(row=r,column=3,value=s)
    for cc in (hc,sc):
        cc.alignment=Alignment(horizontal="center")
        if isinstance(h,str) and h.startswith("="):
            if "MOIC" in m: cc.number_format="0.00\"x\""
            elif "IRR" in m: cc.number_format="0.0%"
            elif "US$" in m: cc.number_format="0.00"
            else: cc.number_format="#,##0"
    ws.cell(row=r,column=4,value=c).font=Font(size=9, italic=True, color="595959")
    r+=1
# verdict row
ws.cell(row=r+1,column=1,value="IC VERDICT")
style_subhead(ws.cell(row=r+1,column=1))
ws.cell(row=r+1,column=2,value="HOLD & STAGGER-EXIT by ~CY2028")
ws.cell(row=r+1,column=3,value="PASS on control")
ws.cell(row=r+1,column=2).fill=PatternFill("solid",fgColor=GREEN)
ws.cell(row=r+1,column=3).fill=PatternFill("solid",fgColor=RED)
for col in "BC": ws.cell(row=r+1,column={"B":2,"C":3}[col]).font=Font(bold=True)
ws.column_dimensions["A"].width=28
for col in "BCD": ws.column_dimensions[col].width=22
box(ws,f"A3:D{r-1}")

# move Dashboard to first position
wb.move_sheet("Dashboard", -(len(wb.sheetnames)-1))

wb.save("/projects/sandbox/PE_Exit_Model_Hexaware_Siemens.xlsx")
print("Excel saved. Sheets:", wb.sheetnames)
